"""
Glosario de variables. Lee el archivo institucional
`2026_09_01_Glosario_variables - Aplicativo.xlsx` (hoja Explicacion_Variables)
y lo expone estructurado: categoría, descripción, fuentes y en qué partes del
aplicativo se usa cada variable (filtro, vista previa, descarga).
"""
from __future__ import annotations

from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.config import (
    COLUMN_SECTIONS,
    FILTERS_BY_KEY,
    GLOSARIO_FECHA,
    PREVIEW_COLUMNS,
    QUERY_COLUMNS,
)


RESOURCE_DIR = Path(__file__).resolve().parent / "resources"
GLOSSARY_PATH = RESOURCE_DIR / "2026_09_01_Glosario_variables_Aplicativo.xlsx"
METHODOLOGY_PATH = RESOURCE_DIR / "Metodologia_Tejido_Empresarial.docx"

# Filtros cuyo valor coincide con una variable del resultado (para enlazar glosario ↔ filtros).
FILTER_LABEL_BY_COLUMN = {
    "Departamento de la empresa": "DEPARTAMENTO",
    "Municipio de la empresa": "MUNICIPIO",
    "Tamaño de la empresa": "TAMANO",
    "Rango de antigüedad de la empresa (años)": "RANGO_ANTIGUEDAD",
    "Rango de ingresos operacionales (COP)": "RANGO_INGRESOS",
    "Inversión extranjera": "INVERSION_EXTRANJERA",
    "Código CIIU Rev 4 - Actividad principal": "COD_CIIU_1",
    "Descripción CIIU Rev 4 - Actividad principal": "DESCRIPCION_CIIU_1",
    "Cadena CIIU Rev 4 - Actividad principal": "CADENA_CIIU_1",
    "Valor Agregado - Actividad principal": "VALOR_AGREGADO_CIIU_1",
    "Cadena de segmentación": "CADENA_SEGMENTACION",
    "Trayectoria exportadora": "TRAYECTORIA_EXPORTADORA",
    "¿La empresa ha exportado?": "HA_EXPORTADO",
}


# Variables derivadas que el aplicativo entrega pero que el glosario institucional aún no
# describe. Se documentan aquí como «definición complementaria» sin sustituir al glosario.
SUPPLEMENTARY_DEFINITIONS: dict[str, dict[str, str]] = {
    "Rango de antigüedad de la empresa (años)": {
        "description": "Rango en el que se ubica la antigüedad de la empresa (años desde su constitución a la fecha de corte). Es la versión agrupada de la variable «Antigüedad de la empresa (años)» y se usa como filtro.",
        "sources": "Cálculos ProColombia a partir de RUES y Supersociedades (misma fuente de «Antigüedad de la empresa (años)»).",
    },
    "Rango de ingresos operacionales (COP)": {
        "description": "Rango en el que se ubican los ingresos operacionales de la empresa en pesos colombianos. Es la versión agrupada de la variable «Ingresos operacionales (COP)» y se usa como filtro.",
        "sources": "Cálculos ProColombia a partir de Supersociedades y RUES (misma fuente de «Ingresos operacionales (COP)»).",
    },
}


def section_for(variable: str) -> str:
    for section, columns in COLUMN_SECTIONS:
        if variable in columns:
            return section
    return "Otras variables"


def _split_paragraphs(text: str) -> list[str]:
    return [line.strip() for line in str(text or "").splitlines() if line.strip()]


@lru_cache(maxsize=1)
def load_glossary() -> dict[str, Any]:
    workbook = load_workbook(GLOSSARY_PATH, read_only=True, data_only=True)
    sheet = workbook["Explicacion_Variables"]
    entries: list[dict[str, Any]] = []
    result_labels = list(QUERY_COLUMNS.values())
    result_set = set(result_labels)
    header_seen = False
    for variable, description, sources in sheet.iter_rows(min_row=1, max_col=3, values_only=True):
        if not variable:
            continue
        name = str(variable).strip()
        if not header_seen:
            # Se omiten el título del archivo y la fila de encabezados («Variable»).
            if name.casefold() == "variable":
                header_seen = True
            continue
        filter_key = FILTER_LABEL_BY_COLUMN.get(name)
        entries.append({
            "variable": name,
            "description": str(description or "").strip(),
            "description_paragraphs": _split_paragraphs(description or ""),
            "sources": str(sources or "").strip(),
            "category": section_for(name),
            "in_export": name in result_set,
            "in_preview": name in PREVIEW_COLUMNS,
            "filter_key": filter_key,
            "filter_label": FILTERS_BY_KEY[filter_key]["label"] if filter_key else None,
            "origin": "glosario",
        })
    workbook.close()
    glossary_names = {entry["variable"] for entry in entries}
    for name, definition in SUPPLEMENTARY_DEFINITIONS.items():
        if name in glossary_names or name not in result_set:
            continue
        filter_key = FILTER_LABEL_BY_COLUMN.get(name)
        anchor = name.replace("Rango de ", "")
        position = next((i for i, e in enumerate(entries) if e["variable"].casefold().startswith(anchor[:12].casefold())), len(entries))
        entries.insert(position, {
            "variable": name,
            "description": definition["description"],
            "description_paragraphs": [definition["description"]],
            "sources": definition["sources"],
            "category": section_for(name),
            "in_export": True,
            "in_preview": name in PREVIEW_COLUMNS,
            "filter_key": filter_key,
            "filter_label": FILTERS_BY_KEY[filter_key]["label"] if filter_key else None,
            "origin": "aplicativo",
        })
        glossary_names.add(name)
    missing = [label for label in result_labels if label not in glossary_names]
    categories = [section for section, _ in COLUMN_SECTIONS if any(e["category"] == section for e in entries)]
    if any(e["category"] == "Otras variables" for e in entries):
        categories.append("Otras variables")
    return {
        "entries": entries,
        "count": len(entries),
        "institutional_count": sum(1 for e in entries if e["origin"] == "glosario"),
        "supplementary_count": sum(1 for e in entries if e["origin"] == "aplicativo"),
        "categories": categories,
        "coverage": {
            "export_columns": len(result_labels),
            "defined_export_columns": len(result_labels) - len(missing),
            "missing": missing,
        },
        "updated_at": GLOSARIO_FECHA,
        "file_name": "2026_09_01_Glosario_variables - Aplicativo.xlsx",
    }
