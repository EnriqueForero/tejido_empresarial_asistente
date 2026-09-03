"""
Genera una vista previa HTML de un libro Excel (colores, tipografías, anchos y
paneles congelados aproximados) para revisar el formato sin abrir Excel.

Uso:
    python scripts/vista_previa_excel.py LIBRO.xlsx SALIDA.html [--filas 30] [--columnas 18]
"""
from __future__ import annotations

import argparse
import html
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _color(rgb: str | None, por_defecto: str) -> str:
    if not rgb or not isinstance(rgb, str) or len(rgb) < 6:
        return por_defecto
    return f"#{rgb[-6:]}"


def _formatear(valor, formato: str) -> str:
    if valor is None:
        return ""
    if isinstance(valor, (int, float)) and formato and formato != "General" and formato != "@":
        decimales = 2 if "0.00" in formato else 1 if "0.0" in formato and "0.00" not in formato else 0
        texto = f"{valor:,.{decimales}f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return texto
    return str(valor)


def render(ruta: Path, salida: Path, filas: int, columnas: int) -> None:
    libro = load_workbook(ruta, data_only=True)
    partes = [
        "<meta charset='utf-8'><style>body{font-family:'Maven Pro',Segoe UI,sans-serif;background:#e9edf1;margin:0;padding:24px;color:#0b2233}"
        "h1{font-family:Jost,'Maven Pro',sans-serif;font-size:20px;margin:0 0 6px}.meta{color:#52667a;font-size:12px;margin-bottom:18px}"
        ".hoja{background:#fff;border:1px solid #d5dee5;border-radius:10px;margin-bottom:26px;overflow:auto;box-shadow:0 8px 24px -18px rgba(1,22,39,.5)}"
        ".pestana{display:inline-flex;align-items:center;gap:8px;margin:12px 14px 8px;font:600 13px Jost,sans-serif}.pestana i{width:10px;height:10px;border-radius:2px;background:#ffa400;display:inline-block}"
        "table{border-collapse:collapse;font-size:11px}td{padding:3px 6px;border:1px solid #e1e8ed;vertical-align:top;white-space:pre-wrap;overflow:hidden}"
        ".cabecera-col td{background:#f1f4f7;color:#7a8a99;text-align:center;font-family:'IBM Plex Mono',monospace;font-size:10px;border-color:#d5dee5}"
        ".num-fila{background:#f1f4f7;color:#7a8a99;text-align:center;font-family:'IBM Plex Mono',monospace;font-size:10px;border-color:#d5dee5}</style>",
        f"<h1>{html.escape(ruta.name)}</h1><div class='meta'>Vista previa generada con openpyxl · hojas: {', '.join(libro.sheetnames)}</div>",
    ]
    for hoja in libro.worksheets:
        max_col = min(hoja.max_column, columnas)
        max_fila = min(hoja.max_row, filas)
        anchos = []
        for c in range(1, max_col + 1):
            dim = hoja.column_dimensions.get(get_column_letter(c))
            ancho = dim.width if dim and dim.width else 9
            anchos.append(int(ancho * 7.2))
        combinadas = {}
        for rango in hoja.merged_cells.ranges:
            combinadas[(rango.min_row, rango.min_col)] = (rango.max_row - rango.min_row + 1, rango.max_col - rango.min_col + 1)
            for r in range(rango.min_row, rango.max_row + 1):
                for c in range(rango.min_col, rango.max_col + 1):
                    if (r, c) != (rango.min_row, rango.min_col):
                        combinadas[(r, c)] = None
        partes.append(f"<div class='hoja'><div class='pestana'><i></i>{html.escape(hoja.title)} <span style='color:#52667a;font-weight:400'>· {hoja.dimensions} · paneles: {hoja.freeze_panes or '—'} · filtro: {hoja.auto_filter.ref or '—'}</span></div><table>")
        partes.append("<tr class='cabecera-col'><td></td>" + "".join(f"<td style='width:{anchos[c-1]}px;min-width:{anchos[c-1]}px'>{get_column_letter(c)}</td>" for c in range(1, max_col + 1)) + "</tr>")
        for r in range(1, max_fila + 1):
            alto = hoja.row_dimensions[r].height
            estilo_fila = f" style='height:{int(alto * 1.33)}px'" if alto else ""
            celdas = [f"<td class='num-fila'>{r}</td>"]
            for c in range(1, max_col + 1):
                info = combinadas.get((r, c), "libre")
                if info is None:
                    continue
                celda = hoja.cell(row=r, column=c)
                relleno = celda.fill.fgColor.rgb if celda.fill and celda.fill.fill_type == "solid" else None
                fondo = _color(relleno if isinstance(relleno, str) else None, "#ffffff")
                fuente = celda.font
                color = _color(fuente.color.rgb if fuente and fuente.color and isinstance(fuente.color.rgb, str) else None, "#0b2233")
                estilos = [f"background:{fondo}", f"color:{color}", f"font-family:{fuente.name or 'Maven Pro'},sans-serif", f"font-size:{(fuente.sz or 10) * 1.15:.0f}px"]
                if fuente.b:
                    estilos.append("font-weight:700")
                if fuente.i:
                    estilos.append("font-style:italic")
                if celda.alignment and celda.alignment.horizontal:
                    estilos.append(f"text-align:{celda.alignment.horizontal}")
                elif isinstance(celda.value, (int, float)):
                    estilos.append("text-align:right")
                if celda.alignment and celda.alignment.vertical == "center":
                    estilos.append("vertical-align:middle")
                if celda.border and celda.border.bottom and celda.border.bottom.color and isinstance(celda.border.bottom.color.rgb, str) and celda.border.bottom.style in {"thick", "medium"}:
                    estilos.append(f"border-bottom:3px solid {_color(celda.border.bottom.color.rgb, '#ffa400')}")
                atributos = ""
                if info != "libre" and info is not None:
                    filas_c, cols_c = info
                    atributos = f" rowspan='{filas_c}' colspan='{cols_c}'"
                texto = html.escape(_formatear(celda.value, celda.number_format))
                celdas.append(f"<td{atributos} style='{';'.join(estilos)}'>{texto}</td>")
            partes.append(f"<tr{estilo_fila}>" + "".join(celdas) + "</tr>")
        if hoja.max_row > max_fila or hoja.max_column > max_col:
            partes.append(f"<tr><td colspan='{max_col + 1}' style='color:#52667a;font-style:italic'>… {hoja.max_row} filas × {hoja.max_column} columnas en total</td></tr>")
        partes.append("</table></div>")
    salida.write_text("".join(partes), encoding="utf-8")
    print(f"Vista previa: {salida}")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("libro", type=Path)
    parser.add_argument("salida", type=Path)
    parser.add_argument("--filas", type=int, default=30)
    parser.add_argument("--columnas", type=int, default=18)
    args = parser.parse_args()
    render(args.libro, args.salida, args.filas, args.columnas)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
