"""
Pruebas del asistente de análisis.

Cubren lo que hace que las respuestas se puedan mostrar sin pedirle confianza al
usuario: qué SQL se deja ejecutar, qué cifras se dejan afirmar, qué gráfica
corresponde a cada resultado y qué llevan los archivos que se descargan.

No tocan Snowflake: Cortex Analyst y la sesión se sustituyen por dobles.
"""
from __future__ import annotations

import os

os.environ["APP_DEMO_MODE"] = "true"
os.environ["APP_ENV"] = "development"

from typing import Any

import pandas as pd
import pytest
from fastapi.testclient import TestClient

from backend.config import ALLOWED_SCHEMAS, IA_ADVERTENCIA
from backend.ia import graficos
from backend.ia.analyst import RespuestaAnalyst, cuenta_para_jwt, host_rest, parsear_respuesta
from backend.ia.guardas import validar_sql, verificar_cifras

TABLA = "APP_SEGMENTACION_EXPORTACIONES.SEGMENTACION.TEJIDO_EMPRESARIAL_COMPLETO_BASE_MUNICIPIOS_P"


# ── Guardas: qué SQL se ejecuta ──────────────────────────────────────────


def test_solo_se_ejecutan_consultas_de_lectura() -> None:
    for peligrosa in (
        f"DELETE FROM {TABLA}",
        f"SELECT 1 FROM {TABLA}; DROP TABLE {TABLA}",
        f"UPDATE {TABLA} SET NIT = '1'",
        "CREATE TABLE X AS SELECT 1",
    ):
        veredicto = validar_sql(peligrosa, ALLOWED_SCHEMAS, 100)
        assert veredicto.ok is False, peligrosa
        assert veredicto.motivo


def test_se_rechaza_una_consulta_a_un_esquema_ajeno() -> None:
    veredicto = validar_sql("SELECT * FROM OTRA_BASE.OTRO_ESQUEMA.TABLA", ALLOWED_SCHEMAS, 100)
    assert veredicto.ok is False
    assert "no autorizado" in veredicto.motivo


def test_se_impone_un_tope_de_filas_si_la_consulta_no_lo_trae() -> None:
    veredicto = validar_sql(f"SELECT NIT FROM {TABLA}", ALLOWED_SCHEMAS, 250)
    assert veredicto.ok is True
    assert "LIMIT 250" in veredicto.sql


def test_se_respeta_el_limite_que_ya_venia_en_la_consulta() -> None:
    veredicto = validar_sql(f"SELECT NIT FROM {TABLA} LIMIT 10", ALLOWED_SCHEMAS, 250)
    assert veredicto.ok is True
    assert veredicto.sql.count("LIMIT") == 1
    assert "LIMIT 10" in veredicto.sql


def test_un_comentario_no_puede_esconder_una_instruccion_prohibida() -> None:
    veredicto = validar_sql(f"SELECT 1 FROM {TABLA} /* DROP TABLE X */ LIMIT 1", ALLOWED_SCHEMAS, 10)
    assert veredicto.ok is True  # el comentario se elimina, no se ejecuta nada de él


# ── Guardas: qué cifras se afirman ───────────────────────────────────────


def test_una_cifra_inventada_se_detecta() -> None:
    filas = [["Antioquia", 231544], ["Valle del Cauca", 148903]]
    verificacion = verificar_cifras("Antioquia tiene 231.544 empresas y Caldas 45.912.", filas, 2)
    assert verificacion.ok is False
    assert "45.912" in verificacion.huerfanas


def test_las_cifras_de_la_tabla_se_aceptan_aunque_esten_redondeadas() -> None:
    filas = [["Bogotá, D.C.", 402118.4]]
    verificacion = verificar_cifras("Bogotá suma 402.118 empresas.", filas, 1)
    assert verificacion.ok is True


def test_se_acepta_una_cifra_expresada_en_millones() -> None:
    filas = [["Total", 48_938_863_957.94]]
    verificacion = verificar_cifras("Las exportaciones sumaron USD 48.939 millones.", filas, 1)
    assert verificacion.ok is True


def test_los_anios_y_los_ordinales_no_cuentan_como_cifras_inventadas() -> None:
    verificacion = verificar_cifras("En 2024 las 10 principales empresas crecieron.", [[1]], 1)
    assert verificacion.ok is True


# ── Gráficas ─────────────────────────────────────────────────────────────


def test_una_dimension_y_una_medida_dan_barras_de_un_solo_tono() -> None:
    espec = graficos.sugerir(["DEPARTAMENTO", "EMPRESAS"], [["Antioquia", 120], ["Bogotá, D.C.", 300]])
    assert espec is not None
    assert espec["tipo"] == "barras"
    assert len(espec["series"]) == 1
    assert espec["series"][0]["color"] == graficos.TONO_UNICO


def test_dos_dimensiones_y_una_medida_dan_barras_apiladas_ordenadas() -> None:
    filas = [
        ["Antioquia", "Micro", 100], ["Antioquia", "Pequeña", 20],
        ["Bogotá, D.C.", "Micro", 300], ["Bogotá, D.C.", "Pequeña", 50],
    ]
    espec = graficos.sugerir(["DEPARTAMENTO", "TAMANO", "EMPRESAS"], filas)
    assert espec is not None
    assert espec["tipo"] == "apiladas"
    assert espec["categorias"] == ["Bogotá, D.C.", "Antioquia"]  # de mayor a menor
    assert [serie["nombre"] for serie in espec["series"]] == ["Micro", "Pequeña"]
    assert espec["series"][0]["valores"] == [300.0, 100.0]


def test_una_sola_cifra_se_muestra_como_indicador_y_no_como_una_barra() -> None:
    espec = graficos.sugerir(["TOTAL"], [[1_678_643]])
    assert espec is not None and espec["tipo"] == "indicador"


def test_las_columnas_por_ano_se_leen_como_serie_de_tiempo() -> None:
    espec = graficos.sugerir(["EXPO_2021", "EXPO_2022", "EXPO_2023"], [[1.0, 2.0, 3.0]])
    assert espec is not None
    assert espec["tipo"] == "lineas"
    assert espec["formato"] == "usd"


def test_un_listado_sin_cifras_no_lleva_grafica() -> None:
    assert graficos.sugerir(["NIT", "RAZON_SOCIAL"], [["900123456", "ACME S.A.S."]]) is None


def test_muchas_categorias_se_recortan_y_se_declara_el_recorte() -> None:
    filas = [[f"Municipio {i}", 100 - i] for i in range(30)]
    espec = graficos.sugerir(["MUNICIPIO", "EMPRESAS"], filas)
    assert espec is not None
    assert len(espec["categorias"]) == graficos.MAX_CATEGORIAS
    assert espec["categorias"][-1] == "Otros"
    assert "Otros" in espec["nota"]


def test_los_colores_salen_de_la_paleta_validada_y_en_orden() -> None:
    filas = [["A", "s1", 1], ["A", "s2", 2], ["B", "s1", 3], ["B", "s2", 4]]
    espec = graficos.sugerir(["EJE", "SERIE", "VALOR"], filas)
    assert espec is not None
    assert [serie["color"] for serie in espec["series"]] == graficos.PALETA_CATEGORICA[:2]


# ── Cliente de Cortex Analyst ────────────────────────────────────────────


def test_la_cuenta_se_normaliza_para_el_token() -> None:
    assert cuenta_para_jwt("my17686.us-east-2.aws") == "MY17686"
    assert cuenta_para_jwt("miorg-micuenta.global") == "MIORG"


def test_el_host_rest_se_deriva_de_la_cuenta(monkeypatch) -> None:
    monkeypatch.delenv("SF_HOST", raising=False)
    assert host_rest("my17686.us-east-2.aws") == "my17686.us-east-2.aws.snowflakecomputing.com"
    monkeypatch.setenv("SF_HOST", "https://especial.snowflakecomputing.com/")
    assert host_rest("cualquiera") == "especial.snowflakecomputing.com"


def test_se_extraen_sql_texto_y_sugerencias_de_la_respuesta() -> None:
    cuerpo = {
        "request_id": "abc",
        "message": {
            "content": [
                {"type": "text", "text": "Interpreté la pregunta así."},
                {"type": "sql", "statement": "SELECT 1"},
                {"type": "suggestions", "suggestions": ["¿Y por departamento?"]},
            ]
        },
        "warnings": [{"message": "columna ambigua"}],
    }
    respuesta = parsear_respuesta(cuerpo)
    assert respuesta.sql == "SELECT 1"
    assert "Interpreté" in respuesta.interpretacion
    assert respuesta.sugerencias == ["¿Y por departamento?"]
    assert respuesta.advertencias == ["columna ambigua"]


# ── Orquestador ──────────────────────────────────────────────────────────


class _ServicioFalso:
    """Doble de SnowflakeService: devuelve una tabla fija y registra lo auditado."""

    def __init__(self, marco: pd.DataFrame | None = None, error: str = "") -> None:
        self._marco = marco if marco is not None else pd.DataFrame({"DEPARTAMENTO": ["Antioquia"], "EMPRESAS": [231544]})
        self._error = error
        self.auditado: list[tuple[str, ...]] = []
        self.consultas: list[str] = []

    def dataframe(self, sql: str) -> pd.DataFrame:
        self.consultas.append(sql)
        if self._error:
            raise RuntimeError(self._error)
        return self._marco

    def filas_con_parametros(self, query: str, parametros: list[Any]) -> list[Any]:
        return [["Antioquia concentra 231.544 empresas."]]

    def log_event(self, *args: str) -> None:
        self.auditado.append(tuple(args))


class _AnalystFalso:
    def __init__(self, sql: str = "", texto: str = "") -> None:
        self._sql = sql
        self._texto = texto
        self.llamadas = 0
        self.vista_semantica = "VISTA"

    def preguntar(self, pregunta: str, historial=None) -> RespuestaAnalyst:
        self.llamadas += 1
        return RespuestaAnalyst(sql=self._sql, interpretacion=self._texto)


def _eventos(servicio, analyst, pregunta="¿Cuántas empresas hay en Antioquia?"):
    from backend.ia.orquestador import Orquestador

    return list(Orquestador(servicio, analyst).procesar(pregunta))


def test_el_flujo_completo_entrega_texto_tabla_grafica_y_advertencia() -> None:
    servicio = _ServicioFalso()
    eventos = _eventos(servicio, _AnalystFalso(sql=f"SELECT DEPARTAMENTO_EMP, COUNT(*) FROM {TABLA} GROUP BY 1"))
    etapas = [evento["etapa"] for evento in eventos if evento["tipo"] == "etapa"]
    final = eventos[-1]

    assert etapas[:3] == ["interpretando", "validando", "consultando"]
    assert final["tipo"] == "final"
    assert "231.544" in final["texto"]
    assert final["columnas"] == ["DEPARTAMENTO", "EMPRESAS"]
    assert final["n_filas"] == 1
    assert final["advertencia"] == IA_ADVERTENCIA
    assert final["meta"]["cifras_verificadas"] is True
    assert servicio.auditado and servicio.auditado[0][0] == "Asistente"


def test_una_sql_peligrosa_del_modelo_nunca_llega_a_la_base() -> None:
    servicio = _ServicioFalso()
    eventos = _eventos(servicio, _AnalystFalso(sql=f"DROP TABLE {TABLA}"))
    assert eventos[-1]["tipo"] == "error"
    assert "seguridad" in eventos[-1]["mensaje"]
    assert servicio.consultas == []  # no se ejecutó nada


def test_si_la_consulta_falla_se_pide_una_correccion_una_sola_vez() -> None:
    servicio = _ServicioFalso(error="SQL compilation error: invalid identifier")
    analyst = _AnalystFalso(sql=f"SELECT MAL FROM {TABLA}")
    eventos = _eventos(servicio, analyst)
    assert any(evento.get("etapa") == "corrigiendo" for evento in eventos)
    assert analyst.llamadas == 2
    assert eventos[-1]["tipo"] == "error"
    assert "no pudo ejecutar" in eventos[-1]["mensaje"]


def test_si_el_modelo_no_produce_sql_se_responde_sin_inventar() -> None:
    eventos = _eventos(_ServicioFalso(), _AnalystFalso(sql="", texto="Necesito más detalle."))
    final = eventos[-1]
    assert final["tipo"] == "final"
    assert final["sql"] == ""
    assert final["filas"] == []
    assert "detalle" in final["texto"]


def test_una_redaccion_con_cifras_sin_respaldo_se_reemplaza_por_los_datos(monkeypatch) -> None:
    servicio = _ServicioFalso()
    monkeypatch.setattr(
        servicio, "filas_con_parametros", lambda *a, **k: [["Antioquia tiene 999.999 empresas."]]
    )
    eventos = _eventos(servicio, _AnalystFalso(sql=f"SELECT DEPARTAMENTO_EMP FROM {TABLA}"))
    final = eventos[-1]
    assert final["meta"]["cifras_verificadas"] is False
    assert "999.999" not in final["texto"]
    assert "231544" in final["texto"] or "231.544" in final["texto"]


def test_una_pregunta_vacia_no_llega_al_servicio() -> None:
    analyst = _AnalystFalso(sql="SELECT 1")
    eventos = _eventos(_ServicioFalso(), analyst, pregunta="   ")
    assert eventos[-1]["tipo"] == "error"
    assert analyst.llamadas == 0


# ── API y descargas ──────────────────────────────────────────────────────


@pytest.fixture()
def cliente() -> TestClient:
    from backend.main import app

    return TestClient(app)


def test_en_modo_demostracion_el_asistente_se_declara_no_disponible(cliente: TestClient) -> None:
    cuerpo = cliente.get("/api/ia/estado").json()
    assert cuerpo["disponible"] is False
    assert "demostración" in cuerpo["motivo"]
    assert cuerpo["advertencia"] == IA_ADVERTENCIA
    assert len(cuerpo["sugerencias"]) >= 8


def test_preguntar_en_modo_demostracion_devuelve_un_error_legible(cliente: TestClient) -> None:
    respuesta = cliente.post("/api/ia/preguntar", json={"pregunta": "¿Cuántas empresas hay?"})
    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"].startswith("text/event-stream")
    assert '"tipo": "error"' in respuesta.text


DATOS_DESCARGA = {
    "pregunta": "¿Cuántas empresas hay por departamento?",
    "respuesta": "Bogotá encabeza con 402.118 empresas.",
    "sql": "SELECT DEPARTAMENTO_EMP, COUNT(DISTINCT NIT) FROM T GROUP BY 1 LIMIT 100",
    "columnas": ["Departamento", "Empresas", "NIT"],
    "filas": [["Bogotá, D.C.", 402118, "900123456"]],
    "n_filas": 1,
}


def test_el_excel_trae_la_pregunta_la_respuesta_la_sql_y_la_advertencia(cliente: TestClient) -> None:
    import io

    import openpyxl

    respuesta = cliente.post("/api/ia/exportar/excel", json=DATOS_DESCARGA)
    assert respuesta.status_code == 200
    assert "attachment" in respuesta.headers["content-disposition"]

    libro = openpyxl.load_workbook(io.BytesIO(respuesta.content))
    assert libro.sheetnames == ["Respuesta", "Datos"]
    textos = " ".join(
        str(celda.value)
        for fila in libro["Respuesta"].iter_rows()
        for celda in fila
        if celda.value is not None
    )
    assert DATOS_DESCARGA["pregunta"] in textos
    assert "inteligencia artificial" in textos
    assert "SELECT DEPARTAMENTO_EMP" in textos
    # Los identificadores conservan su forma: nunca se convierten en número.
    assert libro["Datos"].cell(2, 3).value == "900123456"


def test_la_presentacion_se_genera_con_portada_tabla_y_trazabilidad(cliente: TestClient) -> None:
    import io

    # `python-pptx` es una dependencia de producción (requirements-api.txt). En un
    # entorno mínimo la prueba se omite en lugar de tumbar toda la ejecución.
    Presentation = pytest.importorskip(
        "pptx", reason="requiere python-pptx (viene en requirements-api.txt)"
    ).Presentation

    respuesta = cliente.post("/api/ia/exportar/pptx", json=DATOS_DESCARGA)
    assert respuesta.status_code == 200

    presentacion = Presentation(io.BytesIO(respuesta.content))
    laminas = list(presentacion.slides)
    assert len(laminas) >= 3
    texto = " ".join(
        forma.text_frame.text for lamina in laminas for forma in lamina.shapes if forma.has_text_frame
    )
    assert DATOS_DESCARGA["pregunta"] in texto
    assert "inteligencia artificial" in texto
    assert "SELECT DEPARTAMENTO_EMP" in texto


# ── Tamaño del texto que se le envía al modelo ───────────────────────────


def test_la_tabla_del_prompt_se_acota_por_tamano_no_solo_por_filas() -> None:
    """Una tabla ancha no puede inflar el prompt: el tiempo de redacción crece con él."""
    from backend.ia.redactor import _MAX_CARACTERES_TABLA, tabla_markdown

    columnas = [f"COLUMNA_{i}" for i in range(20)]
    anchas = [[f"valor {fila}-{col} con texto largo de relleno" for col in range(20)] for fila in range(30)]
    texto = tabla_markdown(columnas, anchas, 30)

    assert len(texto) <= _MAX_CARACTERES_TABLA + len(" | ".join(columnas)) * 2 + 200
    assert "filas más, no mostradas" in texto  # declara lo que omitió


def test_una_tabla_angosta_no_pierde_filas() -> None:
    from backend.ia.redactor import _MAX_FILAS_PROMPT, tabla_markdown

    filas = [["Bogotá, D.C.", 402118] for _ in range(30)]
    texto = tabla_markdown(["Departamento", "Empresas"], filas, 30)
    assert len(texto.splitlines()) == _MAX_FILAS_PROMPT + 2  # cabecera + separador
    assert "filas más" not in texto


def test_la_respuesta_informa_el_tiempo_de_cada_etapa() -> None:
    """Sin el desglose no se puede saber por qué una consulta tardó."""
    eventos = _eventos(_ServicioFalso(), _AnalystFalso(sql=f"SELECT DEPARTAMENTO_EMP FROM {TABLA}"))
    meta = eventos[-1]["meta"]
    for clave in ("ms_interpretacion", "ms_consulta", "ms_redaccion", "ms_total"):
        assert clave in meta, clave
        assert isinstance(meta[clave], int)
    assert meta["ms_total"] >= meta["ms_redaccion"]
