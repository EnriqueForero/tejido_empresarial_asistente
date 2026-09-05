"""
Pruebas de la puesta a punto del asistente (3.5.0).

Cubren las reglas que hacen que una respuesta lenta o fallida se vea como lo
que es: un fallo de la redacción cuesta **una** llamada, la degradación llega
con su causa, las descargas salen del servidor con todas las filas, las guardas
resisten los casos adversarios y la telemetría registra todas las salidas.

No tocan Snowflake: Cortex Analyst, la sesión y la telemetría son dobles.
"""
from __future__ import annotations

import os

os.environ["APP_DEMO_MODE"] = "true"
os.environ["APP_ENV"] = "development"

import io
import threading
from typing import Any

import openpyxl
import pandas as pd
import pytest
from fastapi.testclient import TestClient

from backend.config import ALLOWED_SCHEMAS, IA_ADVERTENCIA, NITS_EJEMPLO
from backend.ia import forma, graficos
from backend.ia.analyst import ErrorAnalyst
from backend.ia.guardas import validar_sql, verificar_cifras
from backend.ia.redactor import completar, construir_prompt, es_error_de_firma, redactar
from backend.ia.resultados import AlmacenResultados, ResultadoGuardado
from backend.queries import sql_literal
from dobles import AnalystFalso, ServicioFalso, ServicioTelemetriaFalso, TelemetriaFalsa, correr

TABLA = "APP_SEGMENTACION_EXPORTACIONES.SEGMENTACION.TEJIDO_EMPRESARIAL_COMPLETO_BASE_MUNICIPIOS_P"
VISTA = "APP_SEGMENTACION_EXPORTACIONES.SEGMENTACION.TEJIDO_EMPRESARIAL_SEGMENTACION"


# ── Seguridad: literales SQL ─────────────────────────────────────────────


def test_un_literal_con_barra_invertida_queda_cerrado() -> None:
    """`x\\' OR 1=1 --` cerraba el literal: la barra debe escaparse antes que la comilla."""
    literal = sql_literal("x\\' OR 1=1 --")
    assert literal == "'x\\\\'' OR 1=1 --'"
    # Sin las secuencias de escape no queda ninguna comilla suelta dentro del literal.
    interior = literal[1:-1].replace("\\\\", "").replace("''", "")
    assert "'" not in interior
    assert sql_literal("O'Brien") == "'O''Brien'"
    assert sql_literal("a\\b") == "'a\\\\b'"


# ── Guardas: casos adversarios ───────────────────────────────────────────


def _ok(sql: str, tope: int = 250):
    veredicto = validar_sql(sql, ALLOWED_SCHEMAS, tope)
    assert veredicto.ok is True, veredicto.motivo
    return veredicto.sql


def _rechaza(sql: str, fragmento: str = "") -> None:
    veredicto = validar_sql(sql, ALLOWED_SCHEMAS, 250)
    assert veredicto.ok is False, sql
    assert fragmento.lower() in veredicto.motivo.lower(), veredicto.motivo


def test_una_palabra_prohibida_dentro_de_un_literal_no_rechaza_la_consulta() -> None:
    sql = _ok(f"SELECT NIT FROM {TABLA} WHERE SECTOR = 'SET DROP; DELETE' LIMIT 5")
    assert "LIMIT 5" in sql


def test_un_nombre_de_dos_partes_no_escapa_de_los_esquemas_permitidos() -> None:
    _rechaza("SELECT * FROM SEGUIMIENTO.EVENTOS", "no autorizado")


def test_una_tabla_en_una_lista_separada_por_comas_tambien_se_revisa() -> None:
    _rechaza(f"SELECT * FROM {TABLA} A, SEGUIMIENTO.EVENTOS E WHERE A.NIT = E.FILTROS", "no autorizado")


def test_identifier_y_variables_de_sesion_no_pueden_esconder_una_tabla() -> None:
    _rechaza("SELECT * FROM IDENTIFIER('APP_SEGMENTACION_EXPORTACIONES.SEGUIMIENTO.EVENTOS')", "IDENTIFIER")
    _rechaza("SELECT * FROM TABLE($T)", "")
    _rechaza(f"SELECT SYSTEM$WHOAMI() FROM {TABLA}", "sistema")


def test_una_tabla_sin_calificar_se_rechaza_pero_una_cte_se_acepta() -> None:
    _rechaza("SELECT * FROM EVENTOS", "sin calificar")
    sql = _ok(f"WITH t AS (SELECT NIT FROM {TABLA} LIMIT 10) SELECT * FROM t")
    # El LIMIT de la CTE no cuenta como tope de la consulta: se añade uno al final.
    assert sql.rstrip().endswith("LIMIT 250")


def test_el_tope_se_impone_sin_envolver_la_consulta_y_conserva_el_orden() -> None:
    sql = _ok(f"SELECT NIT FROM {TABLA} ORDER BY NIT DESC LIMIT 999999")
    assert sql.count("LIMIT") == 1
    assert "LIMIT 250" in sql
    assert sql.upper().startswith("SELECT NIT")  # no se envolvió en SELECT * FROM (…)
    assert "ORDER BY NIT DESC" in sql


def test_la_consulta_sobre_la_vista_semantica_se_acepta_y_se_acota() -> None:
    sql = _ok(
        f"SELECT * FROM SEMANTIC_VIEW({VISTA} DIMENSIONS empresas.departamento_emp "
        "METRICS empresas.total_empresas) ORDER BY 2 DESC"
    )
    assert sql.rstrip().endswith("LIMIT 250")
    _rechaza("SELECT * FROM SEMANTIC_VIEW(OTRA.VISTA METRICS x.y)", "no autorizado")


def test_un_parentesis_dentro_de_un_literal_no_descuadra_el_tope() -> None:
    sql = _ok(f"SELECT NIT FROM {TABLA} WHERE RAZON_SOCIAL = 'ACME (B' LIMIT 3")
    assert sql.count("LIMIT") == 1 and "LIMIT 3" in sql


def test_desc_de_order_by_no_se_confunde_con_describe() -> None:
    _ok(f"SELECT NIT FROM {TABLA} ORDER BY INGRESOS_OPERACIONALES DESC LIMIT 3")


# ── Guardas: totales ─────────────────────────────────────────────────────


def test_un_total_correcto_se_acepta_solo_si_el_resultado_esta_completo() -> None:
    filas = [["Antioquia", 100], ["Bogotá, D.C.", 250]]
    assert verificar_cifras("Entre las dos suman 350 empresas.", filas, 2).ok is True
    assert verificar_cifras("El promedio es 175 empresas.", filas, 2).ok is True
    # Con el resultado recortado un total sería parcial y engañoso.
    assert verificar_cifras("Entre las dos suman 350 empresas.", filas, 2, truncado=True).ok is False


# ── Redactor: un fallo cuesta una llamada ────────────────────────────────


def _sesion_que(*resultados_o_errores: Any):
    llamadas: list[str] = []

    def sesion(consulta: str, parametros: list[Any]) -> list[Any]:
        llamadas.append(consulta)
        salida = resultados_o_errores[min(len(llamadas), len(resultados_o_errores)) - 1]
        if isinstance(salida, Exception):
            raise salida
        return salida

    return sesion, llamadas


def test_un_error_del_modelo_no_dispara_la_forma_simple() -> None:
    sesion, llamadas = _sesion_que(RuntimeError("100183 (P0000): Insufficient privileges to use model"))
    with pytest.raises(RuntimeError):
        completar(sesion, "modelo", "prompt")
    assert len(llamadas) == 1


def test_solo_un_error_de_firma_dispara_la_forma_simple_una_sola_vez() -> None:
    assert es_error_de_firma(RuntimeError("SQL compilation error: invalid argument types for function 'COMPLETE'"))
    assert not es_error_de_firma(RuntimeError("Request timed out"))
    sesion, llamadas = _sesion_que(
        RuntimeError("SQL compilation error: invalid argument types"), RuntimeError("otra vez")
    )
    with pytest.raises(RuntimeError):
        completar(sesion, "modelo", "prompt")
    assert len(llamadas) == 2  # con opciones + simple, y nada más


def test_la_redaccion_degradada_trae_su_causa_y_cuesta_una_llamada() -> None:
    sesion, llamadas = _sesion_que(RuntimeError("Model 'x' is not available in region"))
    redaccion = redactar(sesion, "¿Cuántas?", ["DEPARTAMENTO", "EMPRESAS"], [["Antioquia", 5]], 1, False, "m", "abc")
    assert redaccion.degradado is True
    assert redaccion.motivo == "redaccion_fallo"
    assert "not available" in redaccion.error
    assert "Antioquia" in redaccion.texto
    assert len(llamadas) == 1


def test_una_respuesta_vacia_se_declara_como_tal() -> None:
    sesion, llamadas = _sesion_que([[""]])
    redaccion = redactar(sesion, "¿Cuántas?", ["DEPARTAMENTO", "EMPRESAS"], [["Antioquia", 5]], 1, False, "m")
    assert redaccion.degradado is True and redaccion.motivo == "respuesta_vacia"
    assert len(llamadas) == 1


def test_el_prompt_acota_las_palabras_y_prohibe_calcular() -> None:
    prompt = construir_prompt("¿Cuántas?", ["A"], [[1]], 1)
    assert "Máximo 90 palabras" in prompt
    assert "No sumes filas" in prompt
    assert "destaca los primeros lugares" in prompt


# ── Forma del resultado ──────────────────────────────────────────────────


def test_las_columnas_de_contacto_se_reconocen_por_nombre_o_por_senal() -> None:
    assert forma.es_columna_contacto("EMAIL")
    assert forma.es_columna_contacto("Correo electrónico")
    assert forma.es_columna_contacto("CORREO_EMPRESA")
    assert forma.es_columna_contacto("Dirección comercial")
    assert not forma.es_columna_contacto("DEPARTAMENTO_EMP")


def test_las_columnas_conocidas_llevan_su_etiqueta_y_los_alias_se_vuelven_legibles() -> None:
    assert forma.columnas_legibles(["DEPARTAMENTO_EMP", "TOTAL_EMPRESAS", "NIT", "EMPRESAS"]) == [
        "Departamento de la empresa", "Total empresas", "NIT", "Empresas",
    ]
    assert forma.columnas_legibles(["NIT", "NIT"]) == ["NIT", "NIT (2)"]


def test_un_listado_se_detecta_por_la_columna_nit_y_conserva_el_orden() -> None:
    assert forma.nits_del_resultado(["NIT", "RAZON_SOCIAL"], [[890903938.0, "A"], ["811000740", "B"], [890903938, "C"]]) == [
        "890903938", "811000740",
    ]
    assert forma.nits_del_resultado(["DEPARTAMENTO", "EMPRESAS"], [["Antioquia", 5]]) == []
    assert forma.nits_del_resultado(["NIT"], [["no es nit"], ["tampoco"], ["900123456"]]) == []


def test_la_grafica_se_pide_con_palabras_explicitas() -> None:
    for frase in ("Gráfica de empresas por departamento", "grafícame las exportaciones", "muéstrame la evolución 2021-2025", "en barras por tamaño", "tabla y gráfica"):
        assert graficos.pide_grafica(frase), frase
    for frase in ("¿Cuántas empresas hay por departamento?", "Top 10 exportadoras de café", "Lístame esas empresas"):
        assert not graficos.pide_grafica(frase), frase


# ── Almacén de resultados ────────────────────────────────────────────────


def _guardado(consulta_id: str, filas: int = 1, columnas: int = 2) -> ResultadoGuardado:
    return ResultadoGuardado(
        consulta_id=consulta_id, sesion_id="s", pregunta="p", sql="SELECT 1",
        columnas=[f"C{i}" for i in range(columnas)], columnas_tecnicas=[f"C{i}" for i in range(columnas)],
        filas=[[0] * columnas for _ in range(filas)], n_filas=filas, truncado=False,
    )


def test_el_almacen_vence_recorta_y_respeta_el_presupuesto_de_celdas() -> None:
    reloj = [0.0]
    almacen = AlmacenResultados(capacidad=2, vigencia=10, max_celdas=100, reloj=lambda: reloj[0])
    almacen.guardar(_guardado("a"))
    almacen.guardar(_guardado("b"))
    almacen.guardar(_guardado("c"))
    assert almacen.obtener("a") is None and almacen.obtener("c") is not None  # capacidad
    reloj[0] = 11
    assert almacen.obtener("b") is None  # vencido
    almacen.guardar(_guardado("grande", filas=60, columnas=2))  # 120 celdas > presupuesto
    almacen.guardar(_guardado("chico"))
    assert almacen.obtener("grande") is None and almacen.obtener("chico") is not None
    almacen.actualizar("chico", texto="listo")
    assert almacen.obtener("chico").texto == "listo"


# ── Orquestador ──────────────────────────────────────────────────────────


def test_una_redaccion_fallida_se_declara_degradada_con_su_causa() -> None:
    servicio = ServicioFalso(redaccion=RuntimeError("Insufficient privileges to use model"))
    eventos, telemetria, _ = correr(servicio, AnalystFalso(sql=f"SELECT DEPARTAMENTO_EMP FROM {TABLA}"))
    final = eventos[-1]
    assert final["tipo"] == "final"
    assert final["meta"]["degradado"] is True
    assert final["meta"]["motivo_degradacion"] == "redaccion_fallo"
    assert servicio.llamadas_complete == 1  # una llamada, no cuatro
    registro = telemetria.registros[-1]
    assert registro["estado"] == "degradada" and "privileges" in registro["error"]


def test_la_telemetria_registra_todas_las_salidas() -> None:
    sql = f"SELECT DEPARTAMENTO_EMP FROM {TABLA}"
    casos = {
        "exito": (ServicioFalso(), AnalystFalso(sql=sql)),
        "rechazada": (ServicioFalso(), AnalystFalso(sql=f"DROP TABLE {TABLA}")),
        "fallo_sql": (ServicioFalso(error="SQL compilation error"), AnalystFalso(sql=sql)),
        "sin_sql": (ServicioFalso(), AnalystFalso(sql="", texto="Sea más concreto.")),
        "fallo_analyst": (ServicioFalso(), AnalystFalso(error=ErrorAnalyst("Cortex Analyst respondió HTTP 403"))),
    }
    for esperado, (servicio, analyst) in casos.items():
        _, telemetria, _ = correr(servicio, analyst)
        assert len(telemetria.registros) == 1, esperado
        registro = telemetria.registros[0]
        assert registro["estado"] == esperado
        assert registro["ms_total"] >= 0 and registro["consulta_id"]
        assert registro["pregunta"].startswith("¿Cuántas")
    _, telemetria, _ = correr(ServicioFalso(), AnalystFalso(sql=sql), pregunta="   ")
    assert telemetria.registros[0]["estado"] == "pregunta_invalida"


def test_el_desglose_separa_la_correccion_de_la_consulta() -> None:
    servicio = ServicioFalso(error="invalid identifier")
    eventos, telemetria, _ = correr(servicio, AnalystFalso(sql=f"SELECT MAL FROM {TABLA}"))
    assert eventos[-1]["tipo"] == "error"
    registro = telemetria.registros[0]
    assert registro["estado"] == "fallo_sql"
    assert registro["intentos_sql"] == 2
    assert "ms_correccion" in registro


def test_el_servidor_conserva_todas_las_filas_aunque_al_navegador_viajen_menos(monkeypatch) -> None:
    from backend.ia import orquestador as modulo

    monkeypatch.setattr(modulo, "IA_MAX_ROWS_CLIENT", 2)
    marco = pd.DataFrame({"NIT": ["890903938", "811000740", "890912462", "900000001", "900000002"], "RAZON_SOCIAL": list("ABCDE")})
    eventos, _, orquestador = correr(ServicioFalso(marco=marco), AnalystFalso(sql=f"SELECT NIT, RAZON_SOCIAL FROM {TABLA}"))
    resultado = next(evento for evento in eventos if evento["tipo"] == "resultado")
    assert len(resultado["filas"]) == 2 and resultado["truncado"] is True and resultado["n_filas"] == 5
    assert resultado["es_listado"] is True and resultado["n_nits"] == 5
    guardado = orquestador.almacen.obtener(resultado["consulta_id"])
    assert guardado is not None and len(guardado.filas) == 5 and guardado.nits[0] == "890903938"
    assert guardado.texto  # el texto se completó al terminar la redacción


def test_las_columnas_llegan_con_etiqueta_legible_y_sin_contacto_si_asi_se_configura(monkeypatch) -> None:
    from backend.ia import orquestador as modulo

    monkeypatch.setattr(modulo, "EXPORT_INCLUDE_CONTACT_FIELDS", False)
    marco = pd.DataFrame({"DEPARTAMENTO_EMP": ["Antioquia"], "EMAIL": ["a@b.co"], "EMPRESAS": [3]})
    eventos, _, _ = correr(ServicioFalso(marco=marco), AnalystFalso(sql=f"SELECT DEPARTAMENTO_EMP, EMAIL FROM {TABLA}"))
    final = eventos[-1]
    assert final["columnas"] == ["Departamento de la empresa", "Empresas"]
    assert final["filas"] == [["Antioquia", 3]]


def test_la_grafica_solo_se_abre_si_se_pide_o_si_es_un_indicador() -> None:
    marco = pd.DataFrame({"DEPARTAMENTO_EMP": ["Antioquia", "Caldas"], "EMPRESAS": [3, 2]})
    sql = f"SELECT DEPARTAMENTO_EMP, COUNT(*) FROM {TABLA} GROUP BY 1"
    eventos, _, _ = correr(ServicioFalso(marco=marco), AnalystFalso(sql=sql), pregunta="¿Cuántas empresas hay por departamento?")
    assert eventos[-1]["grafica"] is not None and eventos[-1]["mostrar_grafica"] is False
    eventos, _, _ = correr(ServicioFalso(marco=marco), AnalystFalso(sql=sql), pregunta="Gráfica de empresas por departamento")
    assert eventos[-1]["mostrar_grafica"] is True
    eventos, _, _ = correr(ServicioFalso(marco=pd.DataFrame({"TOTAL": [1_678_643]})), AnalystFalso(sql=sql), pregunta="¿Cuántas empresas hay?")
    assert eventos[-1]["grafica"]["tipo"] == "indicador" and eventos[-1]["mostrar_grafica"] is True


def test_si_el_navegador_se_va_no_se_redacta_y_queda_registrado() -> None:
    servicio = ServicioFalso()
    cancelado = threading.Event()
    cancelado.set()
    eventos, telemetria, _ = correr(servicio, AnalystFalso(sql=f"SELECT DEPARTAMENTO_EMP FROM {TABLA}"), cancelado=cancelado)
    assert eventos[-1]["tipo"] == "resultado"  # la tabla sí se produjo; el texto no
    assert servicio.llamadas_complete == 0
    assert telemetria.registros[0]["estado"] == "detenida"


def test_el_historial_se_reconstruye_desde_el_servidor_con_los_consulta_id() -> None:
    analyst = AnalystFalso(sql=f"SELECT DEPARTAMENTO_EMP FROM {TABLA}")
    from backend.ia.orquestador import Orquestador

    orquestador = Orquestador(ServicioFalso(), analyst, telemetria=TelemetriaFalsa())
    primero = list(orquestador.procesar("¿Cuántas pymes hay en Antioquia?"))
    consulta_id = primero[-1]["consulta_id"]
    list(orquestador.procesar("¿Cuántas de esas exportan?", consulta_ids=[consulta_id]))
    historial = analyst.historiales[-1]
    assert [turno["role"] for turno in historial] == ["user", "analyst"]
    assert historial[0]["content"][0]["text"] == "¿Cuántas pymes hay en Antioquia?"
    assert historial[1]["content"][0]["statement"].startswith("SELECT DEPARTAMENTO_EMP")
    # Si el servidor ya no tiene la respuesta, sirve el historial del navegador.
    list(orquestador.procesar("¿Y en Caldas?", consulta_ids=["ffffffffffff"], historial=[
        {"role": "user", "content": [{"type": "text", "text": "anterior"}]},
        {"role": "analyst", "content": [{"type": "sql", "statement": "SELECT 1"}]},
    ]))
    assert analyst.historiales[-1][0]["content"][0]["text"] == "anterior"


# ── Telemetría ───────────────────────────────────────────────────────────


def test_la_telemetria_inserta_con_parametros_y_nunca_rompe() -> None:
    from backend.ia.telemetria import COLUMNAS_CONSULTA, Telemetria

    servicio = ServicioTelemetriaFalso()
    telemetria = Telemetria(servicio, "S.T_CONSULTAS", "S.T_DESCARGAS")
    assert telemetria.registrar_consulta({"consulta_id": "abc", "estado": "exito", "pregunta": "p", "ms_total": 5, "exito": True})
    telemetria.registrar_descarga("abc", "sesion", "excel", 135)
    telemetria.vaciar()
    assert len(servicio.inserciones) == 2
    sentencia, parametros = servicio.inserciones[0]
    assert sentencia.startswith("INSERT INTO S.T_CONSULTAS (") and sentencia.count("?") == len(COLUMNAS_CONSULTA)
    assert len(parametros) == len(COLUMNAS_CONSULTA) and "'" not in sentencia
    assert servicio.inserciones[1][1][3] == "excel"

    rota = Telemetria(ServicioTelemetriaFalso(falla=True), "S.T", "S.D")
    rota.registrar_consulta({"consulta_id": "x"})
    rota.vaciar()
    assert rota.estado()["descartados"] == 1 and "does not exist" in rota.estado()["ultimo_error"]
    assert Telemetria(servicio, "S.T", "S.D", activa=False).registrar_consulta({}) is False


# ── Sesión con Snowflake: reintento sólo si la sesión murió ──────────────


def test_solo_un_error_de_sesion_reabre_la_conexion(monkeypatch) -> None:
    from backend.database import SnowflakeService, es_error_de_sesion

    assert es_error_de_sesion(RuntimeError("Session no longer exists. New login required"))
    assert not es_error_de_sesion(RuntimeError("SQL compilation error: invalid identifier"))

    class _Consulta:
        def __init__(self, fallos: list[str]) -> None:
            self._fallos = fallos

        def collect(self) -> list[Any]:
            if self._fallos:
                raise RuntimeError(self._fallos.pop(0))
            return [[1]]

    class _Sesion:
        def __init__(self, fallos: list[str]) -> None:
            self.fallos = fallos
            self.llamadas = 0

        def sql(self, query: str, params=None) -> _Consulta:
            self.llamadas += 1
            return _Consulta(self.fallos)

    servicio = SnowflakeService()
    resets: list[Any] = []
    monkeypatch.setattr(servicio, "_reset_session", lambda fallida=None: resets.append(fallida))

    generica = _Sesion(["SQL compilation error: invalid identifier"])
    monkeypatch.setattr(servicio, "session", lambda intentos=3: generica)
    with pytest.raises(RuntimeError):
        servicio.filas_con_parametros("SELECT 1", [])
    assert generica.llamadas == 1 and resets == []
    assert "invalid identifier" in (servicio.ultimo_error_consulta or "")

    caida = _Sesion(["Session no longer exists"])
    monkeypatch.setattr(servicio, "session", lambda intentos=3: caida)
    assert servicio.filas_con_parametros("SELECT 1", []) == [[1]]
    assert caida.llamadas == 2 and resets == [caida]
    assert servicio.ultimo_error_consulta is None

    # El modo silencioso no deja rastro en el mensaje al usuario.
    servicio.ultimo_error_consulta = "anterior"
    otra = _Sesion(["Object does not exist"])
    monkeypatch.setattr(servicio, "session", lambda intentos=3: otra)
    with pytest.raises(RuntimeError):
        servicio.filas_con_parametros("INSERT", [], silencioso=True)
    assert servicio.ultimo_error_consulta == "anterior"


# ── API ──────────────────────────────────────────────────────────────────


@pytest.fixture()
def cliente() -> TestClient:
    from backend.main import app

    return TestClient(app)


def _sembrar(consulta_id: str = "abc123abc123", nits: list[str] | None = None) -> None:
    from backend.routers.asistente import orquestador_ia

    orquestador_ia().almacen.guardar(
        ResultadoGuardado(
            consulta_id=consulta_id,
            sesion_id="",
            pregunta="¿Cuántas empresas hay por departamento?",
            sql="SELECT DEPARTAMENTO_EMP, COUNT(DISTINCT NIT) FROM T GROUP BY 1 LIMIT 100",
            columnas=["Departamento", "Empresas", "NIT"],
            columnas_tecnicas=["DEPARTAMENTO_EMP", "EMPRESAS", "NIT"],
            filas=[["Bogotá, D.C.", 402118, "900000001"], ["Antioquia", 231544, "900000003"]],
            n_filas=2,
            truncado=False,
            nits=nits if nits is not None else ["900000001", "900000003"],
            texto="Bogotá encabeza con 402.118 empresas.",
        )
    )


def test_el_estado_del_asistente_trae_los_nit_de_ejemplo_y_la_memoria(cliente: TestClient) -> None:
    cuerpo = cliente.get("/api/ia/estado").json()
    assert len(cuerpo["nit_ejemplo"]) == 3 and cuerpo["memoria_turnos"] >= 1 and cuerpo["resultado_minutos"] >= 1
    assert cliente.get("/api/metadata").json()["nit_examples"] == cuerpo["nit_ejemplo"]
    assert NITS_EJEMPLO == ["890903938", "811000740", "890912462"]


def test_una_pregunta_con_historial_malformado_se_rechaza_antes_de_procesar(cliente: TestClient) -> None:
    malo = cliente.post("/api/ia/preguntar", json={"pregunta": "hola", "historial": [{"role": "system", "content": "x"}]})
    assert malo.status_code == 422
    assert cliente.post("/api/ia/preguntar", json={"pregunta": "hola", "consulta_ids": ["../etc"]}).status_code == 422
    assert cliente.post("/api/ia/exportar/excel", json={"consulta_id": "no-valido"}).status_code == 422


def test_la_descarga_sale_del_servidor_con_todas_las_filas(cliente: TestClient) -> None:
    _sembrar()
    respuesta = cliente.post("/api/ia/exportar/excel", json={"consulta_id": "abc123abc123", "sesion_id": "pestana-1"})
    assert respuesta.status_code == 200
    libro = openpyxl.load_workbook(io.BytesIO(respuesta.content))
    assert libro.sheetnames == ["Respuesta", "Datos"]
    assert libro["Datos"].max_row == 3  # encabezado + 2 filas
    assert libro["Datos"].cell(2, 3).value == "900000001"
    textos = " ".join(str(c.value) for fila in libro["Respuesta"].iter_rows() for c in fila if c.value is not None)
    assert "Bogotá encabeza" in textos and "inteligencia artificial" in textos
    assert cliente.post("/api/ia/exportar/excel", json={"consulta_id": "000000000000"}).status_code == 404


def test_el_listado_del_asistente_sale_con_el_formato_estandar(cliente: TestClient) -> None:
    _sembrar("abc123abc124")
    respuesta = cliente.post("/api/ia/exportar/empresas", json={"consulta_id": "abc123abc124"})
    assert respuesta.status_code == 200, respuesta.text
    nombre = respuesta.headers["X-Export-Filename"]
    assert "Asistente_LoteNIT" in nombre
    libro = openpyxl.load_workbook(io.BytesIO(respuesta.content))
    assert libro.sheetnames == ["Resumen", "Vista_Principal", "Datos_Completos", "Diccionario"]
    resumen = " ".join(str(c.value) for fila in libro["Resumen"].iter_rows() for c in fila if c.value is not None)
    assert "Asistente de análisis" in resumen and "¿Cuántas empresas hay por departamento?" in resumen
    assert IA_ADVERTENCIA[:40] in resumen
    assert libro["Datos_Completos"].max_row >= 3  # los dos NIT sintéticos existen en el modo demostración

    _sembrar("abc123abc125", nits=[])
    assert cliente.post("/api/ia/exportar/empresas", json={"consulta_id": "abc123abc125"}).status_code == 422


def test_la_presentacion_tambien_sale_del_servidor(cliente: TestClient) -> None:
    pytest.importorskip("pptx", reason="requiere python-pptx (viene en requirements-api.txt)")
    _sembrar("abc123abc126")
    respuesta = cliente.post("/api/ia/exportar/pptx", json={"consulta_id": "abc123abc126"})
    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.presentationml")


def test_la_descarga_sin_texto_lo_declara_en_el_archivo(cliente: TestClient) -> None:
    from backend.routers.asistente import orquestador_ia

    _sembrar("abc123abc127")
    orquestador_ia().almacen.actualizar("abc123abc127", texto="")
    respuesta = cliente.post("/api/ia/exportar/excel", json={"consulta_id": "abc123abc127"})
    libro = openpyxl.load_workbook(io.BytesIO(respuesta.content))
    textos = " ".join(str(c.value) for fila in libro["Respuesta"].iter_rows() for c in fila if c.value is not None)
    assert "aún estaba en redacción" in textos


# ── Exportador estándar con notas y advertencia ──────────────────────────


def test_el_exportador_estandar_acepta_notas_y_advertencia() -> None:
    from datetime import datetime

    from backend import demo
    from backend.exporter import create_export, filename_for
    from backend.models import SearchRequest

    solicitud = SearchRequest(mode="batch_nits", nits=["900000001"])
    marco = demo.all_rows(solicitud, 10)
    salida = create_export(marco, solicitud, len(marco), [], generated=datetime(2026, 9, 4, 10, 0), notas=[("Origen", "Asistente")], aviso="AVISO DE PRUEBA")
    libro = openpyxl.load_workbook(salida)
    resumen = " ".join(str(c.value) for fila in libro["Resumen"].iter_rows() for c in fila if c.value is not None)
    assert "Asistente" in resumen and "AVISO DE PRUEBA" in resumen
    assert filename_for(solicitud, 1, datetime(2026, 9, 4, 10, 0), prefijo="Asistente").startswith(
        "ProColombia_TejidoEmpresarial_Asistente_LoteNIT_"
    )
    assert "Asistente" not in filename_for(solicitud, 1, datetime(2026, 9, 4, 10, 0))
