from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from backend.config import PREVIEW_COLUMNS, QUERY_COLUMNS
from backend.demo import DEMO_ROWS
from backend.exporter import _safe_value, create_export, filename_for
from backend.glossary import load_glossary
from backend.models import SearchRequest


def _workbook(frame: pd.DataFrame, request: SearchRequest, total: int):
    output = create_export(frame, request, total, load_glossary()["entries"], generated=datetime(2026, 9, 2, 15, 30))
    return load_workbook(BytesIO(output.getvalue()), data_only=False)


def test_single_company_workbook_structure() -> None:
    request = SearchRequest(mode="nit", term="900000001")
    workbook = _workbook(pd.DataFrame([DEMO_ROWS[0]], columns=list(QUERY_COLUMNS.values())), request, 1)

    assert workbook.sheetnames == ["Resumen", "Ficha_Empresa", "Vista_Principal", "Datos_Completos", "Diccionario"]
    principal = workbook["Vista_Principal"]
    assert [cell.value for cell in principal[6]] == PREVIEW_COLUMNS
    assert principal.freeze_panes == "C7"
    assert principal.auto_filter.ref == f"A6:{principal.cell(row=6, column=len(PREVIEW_COLUMNS)).column_letter}7"
    assert principal["A7"].number_format == "@"
    assert principal["A7"].value == "900000001"
    assert principal.column_dimensions["B"].width >= 24
    assert principal["A1"].fill.fgColor.rgb.endswith("011627")
    assert principal["A1"].font.name == "Jost"

    complete = workbook["Datos_Completos"]
    assert complete.freeze_panes == "D7"
    assert complete.max_column == len(QUERY_COLUMNS)
    assert complete.print_title_rows in {"6:6", "$6:$6"}

    profile = workbook["Ficha_Empresa"]
    labels = [profile.cell(row=row, column=1).value for row in range(6, profile.max_row + 1)]
    assert "Identificación y ubicación" in labels
    assert "Exportaciones por periodo (FOB USD)" in labels
    assert "Contacto y representación" in labels
    assert all(column in labels for column in QUERY_COLUMNS.values())

    glossary = workbook["Diccionario"]
    assert [cell.value for cell in glossary[6]] == ["Variable", "Sección", "Descripción", "Fuentes", "Uso en el aplicativo", "Estado"]
    statuses = {glossary.cell(row=row, column=6).value for row in range(7, glossary.max_row + 1)}
    assert statuses == {"Definición validada en el glosario institucional", "Definición complementaria del aplicativo (rango derivado)"}
    assert "Pendiente de definición" not in statuses


def test_multi_company_workbook_has_no_profile_and_reports_pending_columns() -> None:
    request = SearchRequest(mode="filters", filters={"TAMANO": ["Grande"]})
    frame = pd.DataFrame(DEMO_ROWS[:4], columns=list(QUERY_COLUMNS.values()))
    frame["Variable nueva sin glosario"] = "x"
    workbook = _workbook(frame, request, 4)
    assert "Ficha_Empresa" not in workbook.sheetnames
    glossary = workbook["Diccionario"]
    statuses = [glossary.cell(row=row, column=6).value for row in range(7, glossary.max_row + 1)]
    assert "Pendiente de definición" in statuses
    summary = workbook["Resumen"]
    values = [summary.cell(row=row, column=1).value for row in range(1, summary.max_row + 1)]
    assert "Tamaño de la empresa" in values


def test_descriptive_filenames() -> None:
    stamp = datetime(2026, 9, 2, 15, 30)
    assert filename_for(SearchRequest(mode="filters", filters={"DEPARTAMENTO": ["Valle del Cauca"], "TAMANO": ["Grande"]}), 42, stamp) == (
        "ProColombia_TejidoEmpresarial_Segmentacion_Valle-del-Cauca_y-1-mas_2026-09-02_1530_42-empresas.xlsx"
    )
    assert filename_for(SearchRequest(mode="business_name", term="Café de origen"), 1, stamp) == (
        "ProColombia_TejidoEmpresarial_RazonSocial_Cafe-de-origen_2026-09-02_1530_1-empresa.xlsx"
    )
    assert filename_for(SearchRequest(mode="batch_nits", nits=["900000001", "900000002"]), 2, stamp).startswith(
        "ProColombia_TejidoEmpresarial_LoteNIT_2-NIT_"
    )


def test_excel_formula_and_control_characters_are_neutralized() -> None:
    assert _safe_value('=HYPERLINK("https://example.invalid")').startswith("'=")
    assert _safe_value("texto\x00invalido") == "textoinvalido"
    assert _safe_value("-12.5") == "-12.5"
